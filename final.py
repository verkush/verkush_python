"""
CYS Multi-PDF Requirement Extractor  ▸  Dark-Mode GUI  ▸  v2025-06-11

Features
--------
✔ Extract every CYS-based GUID (CYS-HSM, CYS-SHE, etc.) – even legacy lines with spaces
✔ Skip headers, footers, and table rows
✔ Track & append Cadence columns; highlight text diffs across cadences
✔ Dark-theme Tkinter GUI with progress bar (threaded, so it never freezes)
✔ Excel output: auto-size columns **and wrap text** in all cadence/detail cells
"""

import difflib
import os
import re
import subprocess
import threading
import time
from datetime import datetime

import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import tkinter as tk
from tkinter import END, Listbox, filedialog, messagebox, ttk

# ──────────────────────────────────────────────────────────────────────────────
# Regex patterns ─ now tolerant of spaces / extra hyphens in legacy GUID lines
# ──────────────────────────────────────────────────────────────────────────────
GUID_CORE = r"CYS-[A-Z]+"
# e.g.  CYS-HSM_ac0e60_133   or   CYS-HSM_ ac0e60_133
GUID_PATTERN = re.compile(rf"{GUID_CORE}[_\s-]*[\w-]+", re.IGNORECASE)
ANY_GUID_LINE = re.compile(rf".*GUID:\s*{GUID_CORE}[_\s-]*[\w-]+", re.IGNORECASE)

TABLE_ROW = re.compile(r"^\|.*\|$")
HEADER_FOOTER = re.compile(r"GM Confidential|Page \d+|^\s*\d+\s*$", re.IGNORECASE)
HEADING_WITH_GUID = re.compile(r"^\d+(\.\d+)*\s+.*GUID:", re.IGNORECASE)

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
def is_file_locked(path: str) -> bool:
    """Return True if another process has the file open/locked."""
    try:
        if not os.path.exists(path):
            return False
        os.rename(path, path)  # harmless self-rename
        return False
    except OSError:
        return True


def format_paragraph(lines):
    return " ".join(l.strip() for l in lines if l.strip())


def extract_cadence(text: str) -> str:
    """e.g. page text contains '22.22.142'  →  'Cadence 22.22.142'"""
    m = re.search(r"\b(\d+\.\d+\.\d+)\b", text)
    return f"Cadence {m.group(1)}" if m else "Cadence Unknown"


# ──────────────────────────────────────────────────────────────────────────────
# PDF Parsing
# ──────────────────────────────────────────────────────────────────────────────
def extract_requirements_from(pdf_path):
    """Return (requirement_rows, cadence) for a single PDF."""
    try:
        doc = fitz.open(pdf_path)
        cadence = extract_cadence(doc[0].get_text())
        rows = []

        for page in doc:
            lines = [ln.strip() for ln in page.get_text().split("\n") if ln.strip()]
            i = 0
            while i < len(lines):
                line = lines[i]
                guids = GUID_PATTERN.findall(line)
                if guids:
                    # Grab detail lines that belong to this GUID block
                    details = []
                    j = i + 1
                    while j < len(lines):
                        nxt = lines[j]
                        if GUID_PATTERN.search(nxt):
                            break  # next GUID block starts
                        if (
                            not ANY_GUID_LINE.match(nxt)
                            and not HEADER_FOOTER.search(nxt)
                            and not TABLE_ROW.match(nxt)
                            and not HEADING_WITH_GUID.match(nxt)
                        ):
                            details.append(nxt)
                        j += 1

                    if details:
                        detail_text = format_paragraph(details)
                        info_type = (
                            "Information" if "(information only)" in line.lower() else "Requirement"
                        )
                        for g in guids:
                            # normalise: collapse internal whitespace
                            clean = re.sub(r"\s+", "", g)
                            rows.append([clean, info_type, detail_text, cadence, ""])
                    i = j
                else:
                    i += 1

        return rows, cadence

    except Exception as exc:
        messagebox.showerror(
            "PDF Error", f"Failed to process “{os.path.basename(pdf_path)}”:\n{exc}"
        )
        return [], "Cadence Unknown"


# ──────────────────────────────────────────────────────────────────────────────
# Excel routines
# ──────────────────────────────────────────────────────────────────────────────
def highlight_diff(old: str, new: str) -> str:
    diff = difflib.ndiff(old.split(), new.split())
    return " ".join(
        f"*{w[2:]}*" if w.startswith("+ ") else w[2:] for w in diff if not w.startswith("- ")
    )


def save_to_excel(rows, out_path, updating=False):
    """Write list of rows = [guid, info_type, detail, cadence, service]."""
    while is_file_locked(out_path):
        # Give the user 5 seconds to close the workbook if it's open in Excel.
        if messagebox.askretrycancel(
            "File Locked",
            f"“{os.path.basename(out_path)}” is open in another program.\n"
            "Close it first then click Retry.",
        ):
            time.sleep(1)
            continue
        return

    if updating and os.path.exists(out_path):
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        guid_row = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        header = ["Requirement ID", "Requirement/Information"]
        ws.append(header)
        guid_row = {}

    # Make sure cadence columns and “HSE Service” column exist
    for *_, cadence, _ in rows:
        if cadence not in header:
            header.append(cadence)
            ws.cell(row=1, column=len(header)).value = cadence
    if "HSE Service" not in header:
        header.append("HSE Service")
        ws.cell(row=1, column=len(header)).value = "HSE Service"

    # Styling for header (dark fill, white text)
    dark_fill = PatternFill("solid", fgColor="444444")
    for col_idx, head in enumerate(header, 1):
        cell = ws.cell(1, col_idx)
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)

    # Map for convenience
    col_index = {h: i + 1 for i, h in enumerate(header)}
    service_col = col_index["HSE Service"]

    # Insert or update rows
    for guid, info_type, detail, cadence, service in rows:
        if guid in guid_row:
            r = guid_row[guid]
        else:
            r = ws.max_row + 1
            guid_row[guid] = r
            ws.cell(r, col_index["Requirement ID"]).value = guid
            ws.cell(r, col_index["Requirement/Information"]).value = info_type

        c = col_index[cadence]
        cell = ws.cell(r, c)
        prev = cell.value or ""
        new_val = detail if prev == "" else highlight_diff(prev, detail) if prev != detail else prev
        cell.value = new_val
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Add service if blank
        if ws.cell(r, service_col).value in (None, ""):
            ws.cell(r, service_col).value = service

    # Auto-width columns (capped to 60)
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)

    wb.save(out_path)


def auto_filename(example_pdf):
    try:
        with open(example_pdf, "rb") as f:
            txt = f.read().decode("latin1", "ignore")
            m = re.search(r"CYS-([A-Z]+)", txt)
            tag = m.group(1).lower() if m else "guid"
    except Exception:
        tag = "guid"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"cys-{tag}_{ts}.xlsx"


# ──────────────────────────────────────────────────────────────────────────────
# Tkinter GUI
# ──────────────────────────────────────────────────────────────────────────────
root = tk.Tk()
root.title("CYS Multi-PDF Requirement Extractor")
root.geometry("760x540")
root.configure(bg="#1e1e1e")

# -------- Widgets
file_list = Listbox(
    root, selectmode=tk.MULTIPLE, width=90, bg="#2e2e2e", fg="white", borderwidth=0
)
file_list.pack(padx=12, pady=10, fill="both", expand=False)

BTN_STYLE = dict(bg="#333333", fg="white", activebackground="#444444", activeforeground="white")

tk.Button(root, text="Add PDF Files", **BTN_STYLE, command=lambda: (
    paths := filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")]),
    [file_list.insert(END, p) for p in paths if p not in file_list.get(0, END)]
)).pack(pady=3)

tk.Button(
    root,
    text="Remove Selected",
    **BTN_STYLE,
    command=lambda: [file_list.delete(i) for i in reversed(file_list.curselection())],
).pack(pady=3)

update_var = tk.BooleanVar()
tk.Checkbutton(
    root,
    text="Update Existing Excel (choose file when prompted)",
    variable=update_var,
    bg="#1e1e1e",
    fg="white",
    activebackground="#1e1e1e",
    activeforeground="white",
).pack(pady=4)

progress = ttk.Progressbar(root, orient="horizontal", mode="determinate", length=680)
progress.pack(pady=8)

status = tk.Label(root, text="", bg="#1e1e1e", fg="white")
status.pack(pady=2)


# -------- Worker thread
def run_extraction():
    files = file_list.get(0, END)
    if not files:
        messagebox.showwarning("No Files", "Please add PDF files to process.")
        return

    # Collect rows
    progress["value"] = 0
    progress["maximum"] = len(files)
    all_rows = []

    for idx, pdf in enumerate(files, 1):
        status.config(text=f"Reading “{os.path.basename(pdf)}” ({idx}/{len(files)}) …")
        rows, _ = extract_requirements_from(pdf)
        all_rows.extend(rows)
        progress["value"] = idx
        root.update_idletasks()

    if not all_rows:
        status.config(text="No valid requirements found in selected PDFs.")
        messagebox.showinfo("No Data", "No valid requirements were found.")
        progress["value"] = 0
        return

    if update_var.get():
        xlsx_path = filedialog.askopenfilename(
            title="Choose existing Excel file to update", filetypes=[("Excel files", "*.xlsx")]
        )
        if not xlsx_path:
            status.config(text="Update cancelled.")
            progress["value"] = 0
            return
    else:
        xlsx_path = os.path.join(os.path.dirname(files[0]), auto_filename(files[0]))

    save_to_excel(all_rows, xlsx_path, updating=update_var.get())
    status.config(text=f"✅ Done. Excel {'updated' if update_var.get() else 'created'} → {os.path.basename(xlsx_path)}")

    # Auto-open only for new workbooks
    if not update_var.get():
        subprocess.Popen(["start", "", xlsx_path], shell=True)


tk.Button(
    root,
    text="Extract Requirements",
    **BTN_STYLE,
    command=lambda: threading.Thread(target=run_extraction, daemon=True).start(),
).pack(pady=12)

root.mainloop()