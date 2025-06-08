import fitz  # PyMuPDF
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from collections import defaultdict

def is_table_block(block):
    lines = block.get("lines", [])
    return len(lines) > 2 and all(len(line["spans"]) > 1 for line in lines)

def extract_release_cadence(pdf_path):
    doc = fitz.open(pdf_path)
    first_page = doc[0].get_text()
    match = re.search(r"\b\d{2}\.\d{2}\.\d{3}\b", first_page)
    return match.group(0) if match else "UnknownCadence"

def extract_requirements_from_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    requirements = []

    req_pattern = re.compile(r"GUID:\s*(CYS-HSM[^\n\r/]+(?:\s*/\s*CR\s*\d+))", re.IGNORECASE)
    info_pattern = re.compile(r"\(information only\)", re.IGNORECASE)

    text_blocks = []
    for page in doc:
        blocks = page.get_text("dict")["blocks"]
        for block in blocks:
            if block['type'] == 0 and not is_table_block(block):
                text = ' '.join([span['text'] for line in block['lines'] for span in line['spans']])
                text_blocks.append(text.strip())

    i = 0
    while i < len(text_blocks):
        current_match = req_pattern.search(text_blocks[i])
        if current_match:
            current_id = current_match.group(1).strip()
            is_info = bool(info_pattern.search(text_blocks[i]))

            i += 1
            detail_lines = []

            while i < len(text_blocks):
                next_block = text_blocks[i]
                next_match = req_pattern.search(next_block)
                if next_match:
                    j = i + 1
                    has_details = False
                    while j < len(text_blocks):
                        if req_pattern.search(text_blocks[j]):
                            break
                        if text_blocks[j].strip():
                            has_details = True
                            break
                        j += 1
                    if has_details:
                        break
                    else:
                        detail_lines.append(next_block)
                        i += 1
                        continue
                else:
                    detail_lines.append(next_block)
                    i += 1

            details = ' '.join(detail_lines).strip()
            if details:
                tag = "Information" if is_info else "Requirement"
                requirements.append((current_id, tag, details))
        else:
            i += 1

    return requirements

selected_files = []

def add_files():
    files = filedialog.askopenfilenames(title="Select PDF Files", filetypes=[("PDF Files", "*.pdf")])
    for file in files:
        if file not in selected_files:
            selected_files.append(file)
            listbox.insert(tk.END, os.path.basename(file))

def remove_selected():
    selected_indices = listbox.curselection()
    for index in reversed(selected_indices):
        listbox.delete(index)
        del selected_files[index]

def extract_all():
    if not selected_files:
        messagebox.showwarning("No PDFs Selected", "Please add PDF files first.")
        return

    all_requirements = defaultdict(lambda: {"Requirement/Information": ""})
    cadence_columns = []
    cadence_to_reqs = {}

    for path in selected_files:
        cadence_raw = extract_release_cadence(path)
        cadence_name = f"Cadence {cadence_raw}"
        cadence_columns.append(cadence_name)

        reqs = extract_requirements_from_pdf(path)
        cadence_to_reqs[cadence_name] = {}

        for req_id, tag, detail in reqs:
            if req_id not in all_requirements:
                all_requirements[req_id]["Requirement/Information"] = tag
            cadence_to_reqs[cadence_name][req_id] = detail

    for req_id in all_requirements:
        for cadence in cadence_columns:
            all_requirements[req_id][cadence] = cadence_to_reqs.get(cadence, {}).get(req_id, "")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    save_dir = os.path.join(script_dir, "Extracted_Requirement")
    os.makedirs(save_dir, exist_ok=True)
    save_path = os.path.join(save_dir, "All_Requirements.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    headers = ["Requirement ID", "Requirement/Information"] + cadence_columns + ["HSE Service"]
    ws.append(headers)

    highlight_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    for req_id, content in all_requirements.items():
        row = [req_id, content["Requirement/Information"]]
        detail_values = [content.get(c, "") for c in cadence_columns]

        row_idx = ws.max_row + 1
        ws.append(row + detail_values + [""])

        filtered = [v for v in detail_values if v.strip()]
        if len(set(filtered)) > 1:
            for col_idx, value in enumerate(detail_values, start=3):
                if value.strip():
                    ws.cell(row=row_idx, column=col_idx).fill = highlight_fill

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True)

    wb.save(save_path)
    messagebox.showinfo("Success", f"Saved to:\n{save_path}")

root = tk.Tk()
root.title("Multi-PDF Requirement Extractor")
root.geometry("620x450")
root.resizable(False, False)

frame = tk.Frame(root)
frame.pack(pady=10)

label = tk.Label(frame, text="Step 1: Add PDF files below", font=("Arial", 12))
label.pack(pady=5)

listbox = tk.Listbox(frame, width=80, height=10, selectmode=tk.MULTIPLE)
listbox.pack(pady=5)

btn_frame = tk.Frame(root)
btn_frame.pack(pady=10)

add_btn = tk.Button(btn_frame, text="Add PDF", command=add_files, width=20)
add_btn.grid(row=0, column=0, padx=10)

remove_btn = tk.Button(btn_frame, text="Remove Selected", command=remove_selected, width=20)
remove_btn.grid(row=0, column=1, padx=10)

extract_btn = tk.Button(root, text="Extract Requirements", command=extract_all, width=30, bg="#4CAF50", fg="white")
extract_btn.pack(pady=20)

root.mainloop()