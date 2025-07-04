import fitz  # PyMuPDF
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from collections import defaultdict
from datetime import datetime


def format_paragraph(lines):
    return ' '.join(line.strip() for line in lines if line.strip())


def extract_release_cadence(pdf_path):
    doc = fitz.open(pdf_path)
    first_page = doc[0].get_text()
    match = re.search(r"\b\d{2}\.\d{2}\.\d{3}\b", first_page)
    return match.group(0) if match else "UnknownCadence"


def extract_requirements_final(pdf_path):
    doc = fitz.open(pdf_path)
    requirements = []

    header_footer_pattern = re.compile(r"GM Confidential|Page \d+|^\s*\d+\s*$", re.IGNORECASE)
    table_pattern = re.compile(r'^\|.*\|$')
    heading_guid_pattern = re.compile(r"^\d+(\.\d+)*\s+.*GUID:", re.IGNORECASE)
    valid_guid_pattern = re.compile(r"^GUID:\s*CYS-[\w\-]+", re.IGNORECASE)
    any_guid_pattern = re.compile(r".*GUID:\s*CYS-[\w\-]+", re.IGNORECASE)

    for page in doc:
        lines = page.get_text().split('\n')
        lines = [line.strip() for line in lines if line.strip()]

        i = 0
        while i < len(lines):
            line = lines[i]

            if line.lower().startswith("guid:") and i + 1 < len(lines) and lines[i + 1].startswith("CYS-"):
                line = f"{line} {lines[i + 1]}"
                i += 1

            if valid_guid_pattern.match(line):
                j = i + 1
                has_valid_detail = False
                while j < len(lines):
                    next_line = lines[j].strip()
                    if valid_guid_pattern.match(next_line):
                        break
                    if (
                        not any_guid_pattern.match(next_line)
                        and not header_footer_pattern.search(next_line)
                        and not table_pattern.match(next_line)
                        and not heading_guid_pattern.match(next_line)
                    ):
                        has_valid_detail = True
                        break
                    j += 1

                if not has_valid_detail:
                    i = j
                    continue

                match = re.search(r"CYS-[\w\-]+", line)
                req_id_clean = match.group(0) if match else line.strip()
                info_type = "Information" if "(information only)" in line.lower() else "Requirement"

                details = []
                j = i + 1
                while j < len(lines):
                    next_line = lines[j].strip()
                    if valid_guid_pattern.match(next_line):
                        break
                    if (
                        not any_guid_pattern.match(next_line)
                        and not header_footer_pattern.search(next_line)
                        and not table_pattern.match(next_line)
                        and not heading_guid_pattern.match(next_line)
                    ):
                        details.append(next_line)
                    j += 1

                requirements.append([
                    req_id_clean,
                    format_paragraph(details),
                    info_type,
                    ""
                ])
                i = j
            else:
                i += 1

    return requirements


selected_files = []

def add_files():
    files = filedialog.askopenfilenames(title="Select PDF Files", filetypes=[["PDF Files", "*.pdf"]])
    for file in files:
        if file not in selected_files:
            selected_files.append(file)
            listbox.insert(tk.END, os.path.basename(file))

def remove_selected():
    selected_indices = listbox.curselection()
    for index in reversed(selected_indices):
        listbox.delete(index)
        del selected_files[index]

def extract_all():
    if not selected_files:
        messagebox.showwarning("No PDFs Selected", "Please add PDF files first.")
        return

    all_requirements = defaultdict(lambda: {"Requirement/Information": ""})
    cadence_columns = []
    cadence_to_reqs = {}

    for path in selected_files:
        cadence_raw = extract_release_cadence(path)
        cadence_name = f"Cadence {cadence_raw}"
        if cadence_name not in cadence_columns:
            cadence_columns.append(cadence_name)

        reqs = extract_requirements_final(path)
        cadence_to_reqs[cadence_name] = {}

        for req_id, detail, tag, _ in reqs:
            if req_id not in all_requirements:
                all_requirements[req_id]["Requirement/Information"] = tag
            cadence_to_reqs[cadence_name][req_id] = detail

    for req_id in all_requirements:
        for cadence in cadence_columns:
            all_requirements[req_id][cadence] = cadence_to_reqs.get(cadence, {}).get(req_id, "")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    save_dir = os.path.join(script_dir, "Extracted_Requirement")
    os.makedirs(save_dir, exist_ok=True)

    first_guid_match = None
    for path in selected_files:
        text = fitz.open(path)[0].get_text()
        first_guid_match = re.search(r"CYS-[\w\-]+", text)
        if first_guid_match:
            break
    base_name = first_guid_match.group(0) if first_guid_match else "Requirements"
    existing_file_path = None
    for fname in os.listdir(save_dir):
        if fname.startswith(base_name) and fname.endswith(".xlsx"):
            existing_file_path = os.path.join(save_dir, fname)
            break

    append_selected = option_var.get() == "Append to Existing Excel File"
    if append_selected and existing_file_path:
        wb = load_workbook(existing_file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        for cadence_name in cadence_columns:
            if cadence_name not in headers:
                ws.cell(row=1, column=len(headers)+1).value = cadence_name
                headers.append(cadence_name)

        req_id_col = headers.index("Requirement ID") + 1
        req_id_row_map = {ws.cell(row=i, column=req_id_col).value: i for i in range(2, ws.max_row + 1)}

        for cadence_name in cadence_columns:
            cadence_col = headers.index(cadence_name) + 1
            for req_id in all_requirements:
                if req_id in req_id_row_map:
                    detail = all_requirements[req_id].get(cadence_name, "")
                    ws.cell(row=req_id_row_map[req_id], column=cadence_col).value = detail

        wb.save(existing_file_path)
        os.startfile(existing_file_path)
        status_label.config(text=f"Updated existing file: {os.path.basename(existing_file_path)}")
        messagebox.showinfo("Update Complete", f"✅ Cadence data added to existing file:\n{existing_file_path}")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{base_name}_{timestamp}.xlsx"
    save_path = os.path.join(save_dir, excel_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    headers = ["Requirement ID", "Requirement/Information"] + cadence_columns + ["HSE Service"]
    ws.append(headers)

    highlight_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    for req_id in sorted(all_requirements):
        content = all_requirements[req_id]
        row = [req_id, content["Requirement/Information"]]
        detail_values = [content.get(c, "") for c in cadence_columns]

        row_idx = ws.max_row + 1
        ws.append(row + detail_values + [""])

        unique_values = list(set(v.strip() for v in detail_values if v.strip()))

        if len(unique_values) > 1:
            base = unique_values[0]
            for col_idx, value in enumerate(detail_values, start=3):
                if value.strip() and value.strip() != base:
                    ws.cell(row=row_idx, column=col_idx).fill = highlight_fill

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(wrap_text=True)

    col_widths = {
        1: 40,
        2: 22
    }
    for i in range(3, 3 + len(cadence_columns)):
        col_widths[i] = 50
    col_widths[len(headers)] = 20

    for col_idx, width in col_widths.items():
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(64 + (col_idx - 1) % 26 + 1)
        ws.column_dimensions[col_letter].width = width

    wb.save(save_path)
    os.startfile(save_path)
    status_label.config(text=f"Created new file: {os.path.basename(save_path)}")
    messagebox.showinfo("New File Created", f"✅ Extracted {len(all_requirements)} requirements.\nSaved to:\n{save_path}")


root = tk.Tk()
root.title("Multi-PDF Requirement Extractor")
root.geometry("620x520")
root.resizable(False, False)

frame = tk.Frame(root)
frame.pack(pady=10)

label = tk.Label(frame, text="Step 1: Add PDF files below", font=("Arial", 12))
label.pack(pady=5)

listbox = tk.Listbox(frame, width=80, height=10, selectmode=tk.MULTIPLE)
listbox.pack(pady=5)

btn_frame = tk.Frame(root)
btn_frame.pack(pady=10)

add_btn = tk.Button(btn_frame, text="Add PDF", command=add_files, width=20)
add_btn.grid(row=0, column=0, padx=10)

remove_btn = tk.Button(btn_frame, text="Remove Selected", command=remove_selected, width=20)
remove_btn.grid(row=0, column=1, padx=10)

option_frame = tk.Frame(root)
option_frame.pack(pady=5)

option_label = tk.Label(option_frame, text="Step 2: Choose Output Option", font=("Arial", 12))
option_label.pack()

option_var = tk.StringVar(value="Create New Excel File")
option_menu = tk.OptionMenu(option_frame, option_var, "Create New Excel File", "Append to Existing Excel File")
option_menu.config(width=30)
option_menu.pack()

extract_btn = tk.Button(root, text="Extract Requirements", command=extract_all, width=30, bg="#4CAF50", fg="white")
extract_btn.pack(pady=20)

status_label = tk.Label(root, text="", font=("Arial", 11), fg="blue")
status_label.pack()

root.mainloop()
