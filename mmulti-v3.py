
"""
CYS Multi-PDF Requirement Extractor (Optimized with Enhanced GUI)

Features:
---------
✔ Parse multiple PDF files and extract all CYS GUIDs (e.g., CYS-HSM, CYS-SHE, etc.)
✔ Support legacy lines with multiple GUIDs like "Legacy GUID: CYS-HSM_abc / CYS-SHE_xyz"
✔ Extract requirement details, ignoring headers/footers and tables
✔ Dynamically extract and append Cadence number (e.g., Cadence 22.22.142)
✔ Highlight differences in details for same GUID across cadence versions
✔ Append new cadence columns to previously saved Excel files
✔ GUI: Add/Remove PDFs, update existing Excel, display status, show progress bar
✔ Auto save Excel file in script's folder with timestamped name
✔ Automatically open the generated Excel file (for new file creation only)
"""

import fitz  # PyMuPDF
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox, Listbox, END, ttk
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
import difflib
import subprocess
import threading

# Regex Patterns
table_pattern = re.compile(r'^\|.*\|$')
header_footer_pattern = re.compile(r"GM Confidential|Page \d+|^\s*\d+\s*$", re.IGNORECASE)
heading_guid_pattern = re.compile(r"^\d+(\.\d+)*\s+.*GUID:", re.IGNORECASE)
any_guid_pattern = re.compile(r".*GUID:\s*CYS-[A-Z]+[_-]?[\w-]+", re.IGNORECASE)
guid_extract_pattern = re.compile(r"CYS-[A-Z]+[_-]?[\w-]+", re.IGNORECASE)

def format_paragraph(lines):
    return ' '.join(line.strip() for line in lines if line.strip())

def extract_cadence(text):
    match = re.search(r"\b(\d+\.\d+\.\d+)\b", text)
    return f"Cadence {match.group(1)}" if match else "Cadence Unknown"

def extract_requirements_final(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        cadence = extract_cadence(doc[0].get_text())
        requirements = []
        for page in doc:
            lines = [line.strip() for line in page.get_text().split('\n') if line.strip()]
            i = 0
            while i < len(lines):
                line = lines[i]
                guid_matches = guid_extract_pattern.findall(line)
                if guid_matches:
                    details = []
                    j = i + 1
                    while j < len(lines):
                        next_line = lines[j]
                        if guid_extract_pattern.search(next_line):
                            break
                        if not any_guid_pattern.match(next_line) and not header_footer_pattern.search(next_line) and not table_pattern.match(next_line) and not heading_guid_pattern.match(next_line):
                            details.append(next_line)
                        j += 1
                    if details:
                        detail = format_paragraph(details)
                        info_type = "Information" if "(information only)" in line.lower() else "Requirement"
                        for guid in guid_matches:
                            requirements.append([guid, info_type, detail, cadence, ""])
                    i = j
                else:
                    i += 1
        return requirements, cadence
    except Exception as e:
        messagebox.showerror("Error", f"Failed to process {pdf_path}: {str(e)}")
        return [], "Cadence Unknown"

def highlight_differences(old, new):
    diff = list(difflib.ndiff(old.split(), new.split()))
    return ' '.join([f'*{word[2:]}*' if word.startswith('+ ') else word[2:] for word in diff if not word.startswith('- ')])

def save_to_excel(rows, output_path, update_existing=False):
    if update_existing and os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        guid_row = {
            re.sub(r"\s+", "", str(ws.cell(r, 1).value)): r for r in range(2, ws.max_row + 1)
        }
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        header = ["Requirement ID", "Requirement/Information"]
        ws.append(header)
        guid_row = {}

    # Add cadence columns if missing
    cadences_in_rows = sorted(set(row[3] for row in rows))
    for cadence in cadences_in_rows:
        if cadence not in header:
            header.append(cadence)
            ws.cell(row=1, column=len(header)).value = cadence

    if "HSE Service" not in header:
        header.append("HSE Service")
        ws.cell(row=1, column=len(header)).value = "HSE Service"

    # Refresh column index
    col_index = {h: i + 1 for i, h in enumerate(header)}

    for guid, info_type, detail, cadence, service in rows:
        guid_key = re.sub(r"\s+", "", guid)
        if guid_key in guid_row:
            r = guid_row[guid_key]
        else:
            r = ws.max_row + 1
            ws.cell(row=r, column=1).value = guid
            ws.cell(row=r, column=2).value = info_type
            guid_row[guid_key] = r

        col = col_index[cadence]
        existing = ws.cell(row=r, column=col).value or ""
        ws.cell(row=r, column=col).value = highlight_differences(existing, detail) if existing and existing != detail else detail
        ws.cell(row=r, column=col_index["HSE Service"]).value = service

    # Wrap text and adjust column width
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            if cell.row == 1:
                ws.column_dimensions[cell.column_letter].width = 20
            if cell.value:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 5, 60)

    wb.save(output_path)
    if not update_existing:
        subprocess.run(['start', output_path], shell=True)

def generate_excel_filename(first_pdf):
    with open(first_pdf, 'rb') as f:
        content = f.read().decode('latin1', errors='ignore')
        match = re.search(r'CYS-[A-Z]+', content)
        prefix = match.group(0).split('-')[1].lower() if match else "guid"
    return f"cys-{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# GUI Functions
def browse_pdfs():
    files = filedialog.askopenfilenames(filetypes=[("PDF Files", "*.pdf")])
    for file in files:
        if file not in file_listbox.get(0, END):
            file_listbox.insert(END, file)

def remove_selected():
    for i in file_listbox.curselection()[::-1]:
        file_listbox.delete(i)

def extract_thread():
    files = file_listbox.get(0, END)
    if not files:
        messagebox.showwarning("No Files", "Please add PDF files.")
        return

    progress['maximum'] = len(files)
    all_requirements = []

    for i, pdf in enumerate(files, start=1):
        reqs, _ = extract_requirements_final(pdf)
        all_requirements.extend(reqs)
        progress['value'] = i
        status_label.config(text=f"Processing {os.path.basename(pdf)} ({i}/{len(files)})")
        root.update_idletasks()

    if not all_requirements:
        messagebox.showinfo("No Data", "No valid requirements found.")
        return

    if update_var.get():
        output_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    else:
        output_path = os.path.join(os.path.dirname(files[0]), generate_excel_filename(files[0]))

    if output_path:
        save_to_excel(all_requirements, output_path, update_existing=update_var.get())
        status_label.config(text="✅ Excel updated." if update_var.get() else "✅ New Excel created.")

# GUI Setup
root = tk.Tk()
root.title("CYS Multi-PDF Requirement Extractor")
root.geometry("700x500")
root.configure(bg="#1e1e1e")

file_listbox = Listbox(root, selectmode=tk.MULTIPLE, width=80, bg="#2e2e2e", fg="white")
file_listbox.pack(padx=10, pady=10)

btn_style = {'bg': '#333', 'fg': 'white', 'activebackground': '#444', 'activeforeground': 'white'}

tk.Button(root, text="Add PDF Files", command=browse_pdfs, **btn_style).pack(pady=3)
tk.Button(root, text="Remove Selected", command=remove_selected, **btn_style).pack(pady=3)

update_var = tk.BooleanVar()
tk.Checkbutton(root, text="Update Existing Excel", variable=update_var, bg="#1e1e1e", fg="white").pack(pady=2)

tk.Button(root, text="Extract Requirements", command=lambda: threading.Thread(target=extract_thread).start(), **btn_style).pack(pady=10)

progress = ttk.Progressbar(root, orient="horizontal", mode="determinate", length=600)
progress.pack(pady=5)

status_label = tk.Label(root, text="", bg="#1e1e1e", fg="white")
status_label.pack(pady=2)

root.mainloop()
