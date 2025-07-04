import fitz  # PyMuPDF
import pandas as pd
import re
from pathlib import Path
import os
import platform
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# ----------- Paragraph Formatter -----------

def format_paragraph(lines):
    cleaned = [line.strip().rstrip(".") for line in lines if line]
    paragraph = ". ".join(cleaned).strip()
    return paragraph + "." if paragraph else ""

# ----------- Requirement Extractor -----------

def extract_requirements_final(pdf_path):
    doc = fitz.open(pdf_path)
    requirements = []

    header_footer_pattern = re.compile(r"GM Confidential|Page \d+|^\s*\d+\s*$", re.IGNORECASE)
    table_pattern = re.compile(r'^\|.*\|$')
    heading_guid_pattern = re.compile(r"^\d+(\.\d+)*\s+.*GUID:", re.IGNORECASE)
    valid_guid_pattern = re.compile(r"^GUID:\s*CYS-[\w\-]+.*CR\s+\d+", re.IGNORECASE)
    any_guid_pattern = re.compile(r".*GUID:\s*CYS-[\w\-]+", re.IGNORECASE)

    for page in doc:
        lines = page.get_text().split('\n')
        lines = [line.strip() for line in lines if line.strip()]

        i = 0
        while i < len(lines):
            line = lines[i]

            if valid_guid_pattern.match(line):
                req_id_line = line.strip()
                info_type = "Information" if "(information only)" in req_id_line.lower() else "Requirement"

                details = []
                j = i + 1
                while j < len(lines):
                    next_line = lines[j].strip()

                    if valid_guid_pattern.match(next_line):
                        break

                    if (
                        not any_guid_pattern.match(next_line)
                        and not header_footer_pattern.search(next_line)
                        and not table_pattern.match(next_line)
                        and not heading_guid_pattern.match(next_line)
                    ):
                        details.append(next_line)

                    j += 1

                requirements.append([
                    req_id_line,
                    format_paragraph(details),
                    info_type,
                    ""
                ])
                i = j
            else:
                i += 1

    return requirements

# ----------- Excel Saver & Formatter -----------

def save_to_excel(data, pdf_path):
    df = pd.DataFrame(data, columns=["Requirement ID", "Details", "Requirement/Information", "HSE Service"])
    output_dir = Path("Requirements_Extracted")
    output_dir.mkdir(exist_ok=True)
    file_name = Path(pdf_path).stem + ".xlsx"
    output_path = output_dir / file_name

    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    # Bold header
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Wrap text and resize columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 80)

    wb.save(output_path)
    return output_path

# ----------- Open Folder Cross-Platform -----------

def open_folder(path):
    if platform.system() == "Windows":
        os.startfile(path)
    elif platform.system() == "Darwin":  # macOS
        subprocess.Popen(["open", path])
    else:  # Linux
        subprocess.Popen(["xdg-open", path])

# ----------- GUI Logic -----------

def process_pdf():
    pdf_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    if not pdf_path:
        return

    try:
        extracted = extract_requirements_final(pdf_path)
        if not extracted:
            messagebox.showwarning("No Data", "No valid requirements found.")
            return
        output_path = save_to_excel(extracted, pdf_path)
        messagebox.showinfo("Success", f"{len(extracted)} requirements extracted.\n\nExcel saved to:\n{output_path}")
        open_folder(output_path.parent)
    except Exception as e:
        messagebox.showerror("Error", str(e))

# ----------- GUI Layout -----------

root = tk.Tk()
root.title("Requirement Extractor")
root.geometry("400x150")

frame = tk.Frame(root, pady=20)
frame.pack()

label = tk.Label(frame, text="Select a PDF file to extract requirements:")
label.pack(pady=10)

btn = tk.Button(frame, text="Select PDF", command=process_pdf)
btn.pack()

root.mainloop()
